import tkinter as tk
from tkinter import messagebox
import openpyxl
import os

# ===== Excel 读取原始数据 =====
input_file ="/Users/ch/Library/CloudStorage/OneDrive-mail.nwpu.edu.cn/code/java/study/human-eval/input.xlsx"  # 输入文件
wb_input = openpyxl.load_workbook(input_file)
ws_input = wb_input.active

data_rows = []
for row in ws_input.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1]:
        data_rows.append({"category": row[0], "data": row[1]})

# 初始化响应列表
responses = [None] * len(data_rows)  # 与 data_rows 一一对应

# ===== GUI 初始化 =====
root = tk.Tk()
root.title("LLM Content Validity Judgment")
root.geometry("800x400")

current_index = 0

question_label = tk.Label(root, text="", font=("Arial", 20), wraplength=700, justify="center")
question_label.pack(pady=40)

def update_question():
    """刷新显示当前问题"""
    row = data_rows[current_index]
    question_label.config(text=f"【{row['category']}】\n\n{row['data']}")

    # 显示上一次的选择
    highlight_selection()

def highlight_selection():
    """高亮用户已做出的选择"""
    current_response = responses[current_index]
    if current_response == "Yes":
        yes_button.config(relief="sunken", bg="lightgreen")
        no_button.config(relief="raised", bg="SystemButtonFace")
    elif current_response == "No":
        no_button.config(relief="sunken", bg="lightcoral")
        yes_button.config(relief="raised", bg="SystemButtonFace")
    else:
        yes_button.config(relief="raised", bg="SystemButtonFace")
        no_button.config(relief="raised", bg="SystemButtonFace")

def record_response(answer):
    """记录当前题目的响应"""
    global current_index
    responses[current_index] = answer
    if current_index < len(data_rows) - 1:
        current_index += 1
        update_question()
    else:
        save_results()

def go_back():
    """返回上一题"""
    global current_index
    if current_index > 0:
        current_index -= 1
        update_question()
    else:
        messagebox.showinfo("提示", "已经是第一题了。")

def save_results():
    """保存所有答题结果到Excel"""
    output_file = "results.xlsx"
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.append(["Category", "Data", "Response"])

    for i, row in enumerate(data_rows):
        response = responses[i] if responses[i] else "No Answer"
        ws_output.append([row["category"], row["data"], response])

    wb_output.save(output_file)
    messagebox.showinfo("完成", f"所有项目已评估完毕，结果已保存至 {output_file}")
    root.quit()

# ===== 按钮布局 =====
btn_frame = tk.Frame(root)
btn_frame.pack(pady=20)

yes_button = tk.Button(btn_frame, text="✅ 是（可能为真实）", font=("Arial", 16), width=20,
                       command=lambda: record_response("Yes"))
yes_button.grid(row=0, column=0, padx=10)

no_button = tk.Button(btn_frame, text="❌ 否（不可能为真实）", font=("Arial", 16), width=20,
                      command=lambda: record_response("No"))
no_button.grid(row=0, column=1, padx=10)

back_button = tk.Button(root, text="⬅️ 上一题", font=("Arial", 14), command=go_back)
back_button.pack(pady=10)

# ===== 初始化第一题 =====
update_question()

root.mainloop()